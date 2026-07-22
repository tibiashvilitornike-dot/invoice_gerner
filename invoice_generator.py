# -*- coding: utf-8 -*-
"""
Invoice generation logic extracted from app.py.
build_invoice_xlsx(data) accepts the same JSON payload as the old
/generate-invoice endpoint and returns (BytesIO, filename).
"""
import os
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image

PRODUCT_LIBRARY = {
    "panel": {"name": "სამისამართო მართვის პანელი", "cost_usd": 185},
    "smoke_det": {"name": "სამისამართო კვამლის დეტექტორი ძირით D101", "cost_usd": 8},
    "call_point": {"name": "სამისამართო ხელის ღილაკი D135", "cost_usd": 9},
    "siren": {"name": "სამისამართო სირენა მანათობელი D106", "cost_usd": 12},
    "temp_sensor": {"name": "სამისამართო ტემპერატურული სენსორი D102", "cost_usd": 10},
    "module": {"name": "სამისამართო მოდული D119", "cost_usd": 14},
}


def build_invoice_xlsx(data):
    client_name = data.get('client_name', '')
    invoice_number = data.get('invoice_number', '')
    consumables_pct = data.get('consumables_pct', 0)
    items = data.get('items', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.views.sheetView[0].showGridLines = True

    font_main = "Segoe UI"
    font_company = Font(name=font_main, size=10, bold=False, color="4A5568")
    font_invoice_title = Font(name=font_main, size=20, bold=True, color="1A202C")
    font_client_title = Font(name=font_main, size=11, bold=True, color="2D3748")
    font_header = Font(name=font_main, size=11, bold=True, color="FFFFFF")
    font_data = Font(name=font_main, size=11, bold=False, color="2D3748")
    font_total = Font(name=font_main, size=11, bold=True, color="1A202C")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    light_side = Side(style='thin', color='CBD5E1')
    dark_bottom_side = Side(style='double', color='1A202C')

    border_data = Border(left=light_side, right=light_side, top=light_side, bottom=light_side)
    border_total = Border(top=Side(style='thin', color='94A3B8'), bottom=dark_bottom_side)

    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_total = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_input = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    fill_input_zebra = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    currency_format = '#,##0.00 "₾"'
    usd_format = '#,##0.00 "$"'
    number_format_2dp = '#,##0.00'

    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']

    # 1. კომპანიის რეკვიზიტები
    company_info = (
        "შპს გერნერ\n"
        "ს.კ : 406450889\n"
        "მისამართი: ვაზისუბანი I მკ/რ, 4ა\n"
        "ტელ: +995 593 762 894\n"
        "ანგარიში: GE73BG0000000582427460GEL"
    )
    ws.merge_cells("A2:D5")
    ws["A2"] = company_info
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A2"].font = font_company

    for r in range(2, 6):
        ws.row_dimensions[r].height = 18

    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 110
        img.height = 55
        ws.add_image(img, "H2")

    ws["J2"] = "ინვოისი"
    ws["J2"].font = font_invoice_title
    ws["J2"].alignment = align_right

    ws["J3"] = f"№: {invoice_number}"
    ws["J3"].font = Font(name=font_main, size=12, bold=True, color="4A5568")
    ws["J3"].alignment = align_right

    # 2. დამკვეთის ბლოკი
    ws["A8"] = "დამკვეთი:"
    ws["A8"].font = Font(name=font_main, size=10, bold=True, color="64748B")
    ws["A9"] = client_name
    ws["A9"].font = font_client_title

    # 3. პროექტის დასახელება
    ws["A13"] = "მომსახურების დასახელება:"
    ws["A13"].font = Font(name=font_main, size=10, bold=True, color="64748B")
    ws["A14"] = "სახანძრო სიგნალიზაციის მიწოდება და მონტაჟი"
    ws["A14"].font = Font(name=font_main, size=13, bold=True, color="1E293B")
    ws.row_dimensions[14].height = 22

    # 4. ცხრილის ჰედერები
    headers = [
        "დასახელება", "ერთეული", "რაოდენობა", "ღირებულება\n(USD)",
        "კურსი\n(USD→₾)", "ტრანსპ.\n(%)", "მოგება\n(%)",
        "ერთ. ფასი", "მასალის ჯამი", "მონტაჟი", "მონტაჟის ჯამი"
    ]

    ws.row_dimensions[17].height = 32
    for col, header in zip(cols, headers):
        cell = ws[f"{col}17"]
        cell.value = header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = Border(top=light_side, bottom=Side(style='medium', color='1E293B'))
        cell.fill = fill_header

    # 5. პროდუქტების შევსება
    start_row = 18
    current_row = start_row

    for idx, item in enumerate(items):
        ws.row_dimensions[current_row].height = 24
        is_even = (idx % 2 == 0)

        item_id = item.get('product_id')
        is_manual = item.get('manual', False)

        if item_id in PRODUCT_LIBRARY:
            product_name = PRODUCT_LIBRARY[item_id]['name']
            cost_usd = PRODUCT_LIBRARY[item_id]['cost_usd']
        else:
            product_name = item.get('name', '')
            cost_usd = item.get('cost_usd', 0)

        # allow overriding library cost from the request
        if item.get('cost_usd') not in (None, 0):
            cost_usd = item['cost_usd']

        ws[f"A{current_row}"] = product_name
        ws[f"B{current_row}"] = item.get('unit', 'ცალი')
        ws[f"C{current_row}"] = item.get('qty', 0)

        if is_manual:
            ws[f"D{current_row}"] = None
            ws[f"E{current_row}"] = None
            ws[f"F{current_row}"] = None
            ws[f"G{current_row}"] = None
            ws[f"H{current_row}"] = item.get('price', 0)
        else:
            ws[f"D{current_row}"] = cost_usd
            ws[f"E{current_row}"] = item.get('exchange_rate', 0)
            ws[f"F{current_row}"] = item.get('transport_pct', 0)
            ws[f"G{current_row}"] = item.get('profit_pct', 0)
            ws[f"H{current_row}"] = (
                f"=D{current_row}*E{current_row}"
                f"*(1+F{current_row}/100)"
                f"*(1+G{current_row}/100)"
            )

        ws[f"I{current_row}"] = f"=C{current_row}*H{current_row}"
        ws[f"J{current_row}"] = item.get('install_price', 0)
        ws[f"K{current_row}"] = f"=C{current_row}*J{current_row}"

        ws[f"A{current_row}"].alignment = align_left
        ws[f"B{current_row}"].alignment = align_center
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws[f"{col}{current_row}"].alignment = align_right

        ws[f"D{current_row}"].number_format = usd_format
        ws[f"E{current_row}"].number_format = number_format_2dp
        ws[f"F{current_row}"].number_format = '#,##0.0 "%"'
        ws[f"G{current_row}"].number_format = '#,##0.0 "%"'
        for col in ['H', 'I', 'J', 'K']:
            ws[f"{col}{current_row}"].number_format = currency_format

        for col in cols:
            cell = ws[f"{col}{current_row}"]
            cell.font = font_data
            cell.border = border_data

        if is_manual:
            for col in ['D', 'E', 'F', 'G']:
                ws[f"{col}{current_row}"].fill = PatternFill(
                    start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            if is_even:
                for col in ['A', 'B', 'C', 'H', 'I', 'J', 'K']:
                    ws[f"{col}{current_row}"].fill = fill_zebra
        else:
            input_fill = fill_input_zebra if is_even else fill_input
            normal_fill = fill_zebra if is_even else None
            for col in ['D', 'E', 'F', 'G']:
                ws[f"{col}{current_row}"].fill = input_fill
            if normal_fill:
                for col in ['A', 'B', 'C', 'H', 'I', 'J', 'K']:
                    ws[f"{col}{current_row}"].fill = normal_fill

        current_row += 1

    end_data_row = current_row - 1

    # ---- სახარჯი მასალები ----
    ws.row_dimensions[current_row].height = 24
    ws[f"A{current_row}"] = f"სახარჯი მასალები ({consumables_pct}%)"

    if end_data_row >= start_row:
        ws[f"K{current_row}"] = (
            f"=(SUM(I{start_row}:I{end_data_row})"
            f"+SUM(K{start_row}:K{end_data_row}))*{consumables_pct}/100"
        )
    else:
        ws[f"K{current_row}"] = 0

    ws[f"A{current_row}"].alignment = align_left
    ws[f"K{current_row}"].alignment = align_right
    ws[f"K{current_row}"].number_format = currency_format

    for col in cols:
        cell = ws[f"{col}{current_row}"]
        cell.font = font_data
        cell.border = border_data

    consumables_row = current_row
    current_row += 1

    # 6. ჯამი (დღგ-ს გარეშე)
    ws.row_dimensions[current_row].height = 26
    ws[f"A{current_row}"] = "ჯამი (დღგ-ს გარეშე):"
    ws[f"A{current_row}"].font = font_total
    ws[f"A{current_row}"].alignment = align_left

    if end_data_row >= start_row:
        ws[f"K{current_row}"] = (
            f"=SUM(I{start_row}:I{end_data_row})"
            f"+SUM(K{start_row}:K{end_data_row})+K{consumables_row}"
        )
    else:
        ws[f"K{current_row}"] = f"=K{consumables_row}"

    ws[f"K{current_row}"].font = font_total
    ws[f"K{current_row}"].alignment = align_right
    ws[f"K{current_row}"].number_format = currency_format

    for col in cols:
        ws[f"{col}{current_row}"].border = Border(top=Side(style='thin', color='94A3B8'))

    total_row = current_row
    current_row += 1

    # 7. დღგ 18%
    ws.row_dimensions[current_row].height = 24
    ws[f"A{current_row}"] = "დღგ 18%:"
    ws[f"A{current_row}"].font = font_data
    ws[f"A{current_row}"].alignment = align_left

    ws[f"K{current_row}"] = f"=K{total_row}*0.18"
    ws[f"K{current_row}"].font = font_data
    ws[f"K{current_row}"].alignment = align_right
    ws[f"K{current_row}"].number_format = currency_format
    vat_row = current_row
    current_row += 1

    # 8. სულ გადასახდელი
    ws.row_dimensions[current_row].height = 30
    ws[f"A{current_row}"] = "სულ გადასახდელი:"
    ws[f"A{current_row}"].font = Font(name=font_main, size=12, bold=True, color="1E293B")
    ws[f"A{current_row}"].alignment = align_left
    ws[f"A{current_row}"].fill = fill_total

    ws[f"K{current_row}"] = f"=K{total_row}+K{vat_row}"
    ws[f"K{current_row}"].font = Font(name=font_main, size=12, bold=True, color="1E293B")
    ws[f"K{current_row}"].alignment = align_right
    ws[f"K{current_row}"].number_format = currency_format
    ws[f"K{current_row}"].fill = fill_total

    for col in cols:
        cell = ws[f"{col}{current_row}"]
        cell.border = border_total
        if col not in ('A', 'K'):
            cell.fill = fill_total

    # სვეტების ზომები
    widths = {'A': 38, 'B': 10, 'C': 10, 'D': 13, 'E': 10,
              'F': 10, 'G': 10, 'H': 14, 'I': 15, 'J': 13, 'K': 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"{invoice_number or 'invoice'}.xlsx"
    return excel_file, filename
